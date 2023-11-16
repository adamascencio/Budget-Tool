#! Python3
# - Download transaction data from capitalone.com and save to a CSV file

import os, csv
from glob import glob
from pathlib import Path
from openpyxl import Workbook

desktop_path = Path.home() / "Desktop"
download_path = Path.home() / "Downloads"
capital_one_dir = desktop_path / "Capital One"

os.makedirs(capital_one_dir, exist_ok=True)

# Find the most recent CSV file in the Downloads folder
csv_files = glob(os.path.join(download_path, "*.csv"))
csv_files.sort(key=os.path.getmtime, reverse=True)
newest_csv = csv_files[0] if csv_files else None

if newest_csv:
    print(f"Found the most recent CSV file: {newest_csv}")
else:
    print("No CSV files found in the Downloads folder")
    exit()

# Convert csv file to xlsx file
wb = Workbook()
ws = wb.active

with open(newest_csv, "r") as f:
    reader = csv.reader(f, delimiter=",")
    for row in reader:
        ws.append(row)

ws.delete_cols(2, 3)  # Delete Posted Date and Card Number columns

rows_to_delete = []

# First pass: determine which rows to delete
for cell in ws["C"]:
    if not cell.value:
        if ws[f"B{cell.row}"].value == "Payment/Credit":
            rows_to_delete.append(cell.row)
        else:
            cell.value = str(float(ws[f"D{cell.row}"].value) * -1)

# Second pass: delete rows from the bottom up
for row in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row, 1)

ws.delete_cols(4, 1)  # Delete Credit column

wb.save(capital_one_dir / "transactions.xlsx")  # Save wb to the Capital One folder
